"""Excel 解析器——通用版，只做单元格提取，不做语义分析

与 PDF 路线对齐：
  - 遍历所有 sheet（不只读 active）
  - 输出格式：sheet 名 + 每行单元格内容（用 | 分隔）
  - 处理合并单元格
  - 不做行业专用结构化拆分（交给 AI FC）

内存优化：
  - try/finally 确保 wb.close() 释放文件句柄
  - 大文件（>5MB）自动切换 read_only 模式流式读取
  - 行数上限保护（防止百万行 Excel 撑爆内存）
"""

from pathlib import Path

# openpyxl 改为懒导入——在 parse() 中导入，节省模块导入内存

from .base import BaseParser, ParsedDocument


# 大文件阈值：超过此大小切换 read_only 模式
_LARGE_FILE_THRESHOLD = 5 * 1024 * 1024  # 5MB
# 单 sheet 最大行数保护（防止内存溢出）
_MAX_ROWS_PER_SHEET = 50000


class ExcelParser(BaseParser):
    """通用 Excel 解析器——只提取单元格内容，结构化交给 AI"""

    def _build_merge_map(self, ws) -> dict:
        """构建合并单元格映射 {(row, col): (top_row, top_col)}

        read_only 模式下 merged_cells.ranges 不可用，返回空 dict
        """
        merge_map = {}
        if ws.merged_cells is None:
            return merge_map
        try:
            for mr in ws.merged_cells.ranges:
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        merge_map[(r, c)] = (mr.min_row, mr.min_col)
        except Exception:
            # read_only 模式或异常情况，跳过合并单元格处理
            from core.logger import get_logger
            get_logger(__name__).debug("合并单元格枚举失败（跳过合并处理）", exc_info=True)
        return merge_map

    def _val(self, ws, merge_map, r, c):
        """获取单元格实际值（处理合并单元格，返回值指向左上角）"""
        key = (r, c)
        if key in merge_map:
            r, c = merge_map[key]
        v = ws.cell(r, c).value
        return "" if v is None else str(v)

    def _sheet_rows_to_text(self, ws, merge_map, row_start, row_end) -> list:
        """把 sheet 的 [row_start, row_end] 行区间转为文本行列表（不含 [Sheet:] 头）"""
        total_cols = ws.max_column or 0
        lines = []
        for ri in range(row_start, row_end + 1):
            cells = []
            has_data = False
            for ci in range(1, total_cols + 1):
                v = self._val(ws, merge_map, ri, ci)
                cells.append(v)
                if v.strip():
                    has_data = True
            if has_data:
                lines.append(f"  行{ri}: " + " | ".join(cells))
        return lines

    def _sheet_to_text(self, ws, merge_map) -> str:
        """把单个 sheet 转为文本（格式对齐 PDF 页面输出）"""
        total_cols = ws.max_column or 0
        max_row = ws.max_row or 0
        # 行数保护
        if max_row > _MAX_ROWS_PER_SHEET:
            max_row = _MAX_ROWS_PER_SHEET

        lines = [f"[Sheet: {ws.title}] ({ws.max_row}行 x {total_cols}列)"]
        lines.extend(self._sheet_rows_to_text(ws, merge_map, 1, max_row))
        if ws.max_row and ws.max_row > _MAX_ROWS_PER_SHEET:
            lines.append(f"  ...（已截断，仅读取前 {_MAX_ROWS_PER_SHEET} 行，共 {ws.max_row} 行）")
        return chr(10).join(lines)

    def iter_sheet_units(self, ws, merge_map, rows_per_unit: int):
        """把 sheet 切成流单元（tier-1 流式契约）：每单元 ≤ rows_per_unit 行

        - 行数 ≤ rows_per_unit：整个 sheet 为一个单元（输出与 _sheet_to_text 一致）
        - 超过：按行切块，第 2 块起携带首行表头（标注"请勿重复提取"），跨块不断头
        Yields: 单元文本 str
        """
        total_cols = ws.max_column or 0
        max_row = min(ws.max_row or 0, _MAX_ROWS_PER_SHEET)
        if max_row == 0:
            return
        if max_row <= rows_per_unit:
            yield self._sheet_to_text(ws, merge_map)
            return
        header = self._sheet_rows_to_text(ws, merge_map, 1, 1)
        n_chunks = (max_row + rows_per_unit - 1) // rows_per_unit
        for k in range(n_chunks):
            rs = k * rows_per_unit + 1
            re_ = min(max_row, rs + rows_per_unit - 1)
            lines = [f"[Sheet: {ws.title}] (块{k + 1}/{n_chunks}，行{rs}-{re_}，共{ws.max_row}行 x {total_cols}列)"]
            if k > 0 and header:
                lines.append("【表头参考｜请勿重复提取】" + header[0].strip())
            lines.extend(self._sheet_rows_to_text(ws, merge_map, rs, re_))
            yield chr(10).join(lines)

    def parse(self, file_path: str) -> ParsedDocument:
        """解析 Excel 所有 sheet，返回完整文本

        与 PDF 路线对齐：只输出原始单元格文本，结构化提取交给 AI FC。

        内存优化：
        - 大文件（>5MB）自动切换 read_only 模式
        - try/finally 确保文件句柄释放
        - 单 sheet 行数上限保护
        """
        file_size = Path(file_path).stat().st_size
        use_read_only = file_size > _LARGE_FILE_THRESHOLD

        wb = None
        all_lines = []
        sheets_info = []
        # 懒导入 openpyxl（节省模块导入内存）
        import openpyxl
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=use_read_only)
            for ws in wb.worksheets:
                merge_map = self._build_merge_map(ws) if not use_read_only else {}
                sheet_text = self._sheet_to_text(ws, merge_map)
                all_lines.append(sheet_text)
                sheets_info.append({
                    "name": ws.title,
                    "rows": ws.max_row or 0,
                    "cols": ws.max_column or 0,
                })
        except Exception:
            # read_only 模式失败时降级为正常模式
            if use_read_only:
                try:
                    if wb is not None:
                        wb.close()
                except Exception:
                    pass  # 清理/关闭失败不影响主流程（OS 兜底回收）
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
                all_lines = []
                sheets_info = []
                for ws in wb.worksheets:
                    merge_map = self._build_merge_map(ws)
                    sheet_text = self._sheet_to_text(ws, merge_map)
                    all_lines.append(sheet_text)
                    sheets_info.append({
                        "name": ws.title,
                        "rows": ws.max_row or 0,
                        "cols": ws.max_column or 0,
                    })
            else:
                raise
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass  # 清理/关闭失败不影响主流程（OS 兜底回收）

        return ParsedDocument(
            raw_text="\n\n".join(all_lines),
            tables=[],
            structured_tables=[],
            metadata={
                "filename": Path(file_path).name,
                "sheets": sheets_info,
                "total_sheets": len(sheets_info),
                "file_size": file_size,
                "read_only_mode": use_read_only,
            },
        )
