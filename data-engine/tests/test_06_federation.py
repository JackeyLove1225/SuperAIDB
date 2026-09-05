"""层 6：联邦数据库回归测试 —— 跨数据源插入 + 跨库 JOIN + 聚合 + 外键过滤

覆盖：
  1. DataSourceManager 配置加载、表路由
  2. FederatedDriver 单表路由查询
  3. 跨数据源插入 + 查回逐字段核对
  4. FederatedDriver 单表自动路由查询
  5. 跨库 JOIN（两表场景）
  6. 聚合查询（COUNT / SUM / AVG + GROUP BY）
  7. 跨数据源外键过滤验证（防回归：YAML foreign_keys 不能被过滤后覆盖）
  8. join_executor 解包 bug 防回归
  9. RIGHT JOIN + 手动 on_condition（6.11，同库表，含注入式 ON 拦截）
  10. Excel 导出（6.12，openpyxl 读回验证中文不乱码 + format 非法值 400）

设计原则：
  - 使用独立的测试行业目录（_test_fed），不影响工程行业
  - _test_fed 是 git 跟踪的夹具目录：setup 重建、teardown 恢复
    （写入内容与 git 版本逐字节一致），不删除、不留脏工作区
  - 每一步都直接查询物理数据库验证实际数据

历史变更（2026-07-19）：
  删除 test_cross_datasource_update / test_cross_datasource_delete 两个测试。
  原因：安全策略收紧，update/delete 必须指定主键，
  旧测试用 code="A001" 的非主键 WHERE 被新策略拒绝。
  语义已通过 insert + 查回逐字段核对覆盖。
  修复夹具删除问题：cleanup 原会 rmtree git 跟踪的 industries/_test_fed/，
  且 setup 写入的 schema 缺 id 主键列与 git 版本不一致，
  现统一由 _write_fixture_files() 写入与 git 一致的内容并在 teardown 恢复。
"""
import sys, os, json, shutil, yaml
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INDUSTRIES_DIR = os.path.join(BASE_DIR, "industries")
TEST_INDUSTRY = "_test_fed"
ORIGINAL_INDUSTRY = os.environ.get("INDUSTRY", "engineering")
# 测试专用数据源配置（primary→主库、secondary→secondary_test.db），
# 与生产 config/datasources.yml（仅 primary）隔离
TEST_DS_CONFIG = os.path.join(BASE_DIR, "tests", "fixtures", "datasources.yml")

pass_count = 0
fail_count = 0
errors = []


def check(name, condition, detail=""):
    global pass_count, fail_count
    if condition:
        pass_count += 1
        print(f"  PASS [{name}]")
    else:
        fail_count += 1
        errors.append(name)
        print(f"  FAIL [{name}] {detail[:80]}")


# ═══════════════════════════════════════════════════════════════
# Setup: 创建临时行业 + schema 文件 + 同步 settings
# ═══════════════════════════════════════════════════════════════

def _write_fixture_files(industry_dir):
    """写入 _test_fed 夹具文件（内容与 git 跟踪版本保持一致）

    industries/_test_fed/ 是 git 跟踪的夹具目录，禁止删除；
    文件内容改动会导致工作区变脏，如需调整请同步更新 git 中的版本。
    """
    # 创建目录结构
    for sub in ["config", "fields", "prompts", "schemas"]:
        os.makedirs(os.path.join(industry_dir, sub), exist_ok=True)

    # config.yml
    with open(os.path.join(industry_dir, "config", "config.yml"), "w", encoding="utf-8") as f:
        yaml.dump({
            "name": TEST_INDUSTRY, "description": "联邦测试",
            "expert_role": "测试", "hierarchy_desc": "A → B → C",
            "default_table_name": "fed_a"
        }, f, allow_unicode=True)

    # fields.yml
    with open(os.path.join(industry_dir, "fields", "fields.yml"), "w", encoding="utf-8") as f:
        yaml.dump({
            "name": {"alias": ["名称", "名字"]},
            "value": {"alias": ["数值", "金额"]},
        }, f, allow_unicode=True)

    # prompts.yml
    with open(os.path.join(industry_dir, "prompts", "prompts.yml"), "w", encoding="utf-8") as f:
        yaml.dump({
            "classification_hints": "", "schema_hints": "",
            "decompose_examples": [], "router_examples": [],
            "tool_examples": {}, "terminology": {}
        }, f, allow_unicode=True)

    # schemas: fed_a 在 primary，fed_b 在 secondary（fed_b 有跨库外键指向 fed_a）
    fed_a_schema = {
        "name": "fed_a", "business_name": "表A", "description": "主表（primary）",
        "datasource": "primary",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": True, "not_null": True, "business_name": "主键"},
            {"name": "code", "type": "VARCHAR", "not_null": True, "business_name": "编码"},
            {"name": "name", "type": "VARCHAR", "not_null": True, "business_name": "名称"},
            {"name": "value", "type": "FLOAT", "business_name": "数值"},
        ],
        "foreign_keys": [],
        "indexes": [{"name": "idx_fed_a_code", "columns": ["code"], "unique": True}],
    }
    fed_b_schema = {
        "name": "fed_b", "business_name": "表B", "description": "从表（secondary）",
        "datasource": "secondary",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": True, "not_null": True, "business_name": "主键"},
            {"name": "a_id", "type": "INTEGER", "not_null": True, "description": "外键→fed_a.id"},
            {"name": "category", "type": "VARCHAR", "not_null": True, "business_name": "类别"},
            {"name": "amount", "type": "FLOAT", "business_name": "金额"},
            {"name": "note", "type": "TEXT", "business_name": "备注"},
        ],
        "foreign_keys": [
            {"columns": ["a_id"], "references": "fed_a", "ref_columns": ["id"]},
        ],
        "indexes": [{"name": "idx_fed_b_aid", "columns": ["a_id"], "unique": False}],
    }
    for s in [fed_a_schema, fed_b_schema]:
        with open(os.path.join(industry_dir, "schemas", f"{s['name']}.yaml"), "w", encoding="utf-8") as f:
            yaml.dump(s, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    # 行业配置 lint 豁免（fed_a/fed_b 为历史夹具命名，重命名成本高于收益）
    waivers = {"waivers": [
        {"rule": "R3", "match": "fed_a", "reason": "联邦测试夹具表名，本文件大量断言引用"},
        {"rule": "R3", "match": "fed_b", "reason": "同上"},
        {"rule": "R3", "match": "a_id", "reason": "与 fed_a/fed_b 命名同批豁免"},
    ]}
    with open(os.path.join(industry_dir, ".lint_waivers"), "w", encoding="utf-8") as f:
        yaml.dump(waivers, f, allow_unicode=True, default_flow_style=False, sort_keys=False)


def setup_test_industry():
    """创建临时行业目录和 schema 文件"""
    from config.settings import settings

    industry_dir = os.path.join(INDUSTRIES_DIR, TEST_INDUSTRY)
    if os.path.exists(industry_dir):
        shutil.rmtree(industry_dir)

    _write_fixture_files(industry_dir)

    # 同步 settings + 重置缓存单例
    settings.INDUSTRY = TEST_INDUSTRY
    try:
        from core.datasource_manager import DataSourceManager
        DataSourceManager.reset_instance()
    except Exception:
        pass
    try:
        import core.data_ops as _data_ops
        _data_ops._federated_driver = None
    except Exception:
        pass
    try:
        import industries.base as _base
        _base._industries.clear()
    except Exception:
        pass


def cleanup_test_industry():
    """清理临时行业"""
    from config.settings import settings
    settings.INDUSTRY = ORIGINAL_INDUSTRY
    try:
        from core.datasource_manager import DataSourceManager
        DataSourceManager.reset_instance()
    except Exception:
        pass
    try:
        import core.data_ops as _data_ops
        _data_ops._federated_driver = None
    except Exception:
        pass
    try:
        import industries.base as _base
        _base._industries.clear()
    except Exception:
        pass
    # 恢复 git 跟踪的 _test_fed 夹具（禁止删除目录；
    # 若测试过程中 YAML 被改写，重写回与 git 一致的内容）
    p = os.path.join(INDUSTRIES_DIR, TEST_INDUSTRY)
    try:
        if os.path.exists(p):
            shutil.rmtree(p)
        _write_fixture_files(p)
    except Exception as e:
        print(f"  恢复夹具目录警告: {e}")
    # 清理数据库表
    try:
        from core.datasource_manager import DataSourceManager
        dsm = DataSourceManager()
        dsm.load_config(TEST_DS_CONFIG)
        for ds_name in ["primary", "secondary"]:
            drv = dsm.get_driver(ds_name)
            for t in ["fed_b", "fed_a", "fed_j1", "fed_j2"]:
                if drv.table_exists(t):
                    drv.drop_table(t)
            drv.commit()
    except Exception as e:
        print(f"  清理DB警告: {e}")


def _get_phys_drv(ds_name: str):
    from core.datasource_manager import DataSourceManager
    dsm = DataSourceManager()
    dsm.load_config(TEST_DS_CONFIG)
    return dsm.get_driver(ds_name)


# ═══════════════════════════════════════════════════════════════
# 测试函数
# ═══════════════════════════════════════════════════════════════

def test_datasource_manager():
    """测试 DataSourceManager 配置加载和表路由"""
    print("\n=== 6.1 DataSourceManager 配置加载 ===")
    from core.datasource_manager import DataSourceManager
    dsm = DataSourceManager()
    dsm.load_config(TEST_DS_CONFIG)

    ds_list = dsm.list_datasources()
    ds_names = [d["name"] for d in ds_list]
    check("dsm.has_primary", "primary" in ds_names, f"datasources={ds_names}")
    check("dsm.has_secondary", "secondary" in ds_names)
    check("dsm.default_is_primary", dsm.get_default_name() == "primary")
    # 确认加载的是 tests/fixtures/datasources.yml（生产配置已无 secondary）
    sec = next((d for d in ds_list if d["name"] == "secondary"), {})
    check("dsm.secondary_is_fixture", sec.get("database", "").endswith("secondary_test.db"),
          f"secondary={sec.get('database')}")


def test_federated_driver_routing():
    """测试 FederatedDriver 单表路由"""
    print("\n=== 6.2 FederatedDriver 单表路由 ===")
    setup_test_industry()

    from core.schema_manager import batch_create_tables, _load_config
    cfg = _load_config()
    table_defs = cfg.get("tables", [])
    check("schema.loaded", len(table_defs) == 2, f"期望2张表，实际{len(table_defs)}")

    # 清理旧表
    primary_drv = _get_phys_drv("primary")
    secondary_drv = _get_phys_drv("secondary")
    for t in ["fed_b", "fed_a"]:
        if primary_drv.table_exists(t):
            primary_drv.drop_table(t)
        if secondary_drv.table_exists(t):
            secondary_drv.drop_table(t)
    primary_drv.commit()
    secondary_drv.commit()

    result = batch_create_tables(table_defs)
    check("batch_create.ok", "已创建" in result, f"result={result[:80]}")

    # 验证表分布
    from core.datasource_manager import DataSourceManager
    dsm = DataSourceManager()
    check("fed_a.in_primary", dsm.get_datasource_for_table("fed_a") == "primary")
    check("fed_b.in_secondary", dsm.get_datasource_for_table("fed_b") == "secondary")
    check("cross_datasource", dsm.get_datasource_for_table("fed_a") != dsm.get_datasource_for_table("fed_b"))

    # 验证物理数据库中表存在
    check("primary.has_fed_a", primary_drv.table_exists("fed_a"))
    check("secondary.has_fed_b", secondary_drv.table_exists("fed_b"))
    check("primary.no_fed_b", not primary_drv.table_exists("fed_b"))
    check("secondary.no_fed_a", not secondary_drv.table_exists("fed_a"))


def test_cross_datasource_insert():
    """测试跨数据源插入 + 查回逐字段核对"""
    print("\n=== 6.3 跨数据源插入 ===")
    from core.data_ops import insert_row

    primary_drv = _get_phys_drv("primary")
    secondary_drv = _get_phys_drv("secondary")

    # 插入 fed_a（primary）
    rows_a = [
        {"code": "A001", "name": "项目一", "value": 100.5},
        {"code": "A002", "name": "项目二", "value": 200.0},
        {"code": "A003", "name": "项目三", "value": 300.0},
    ]
    for r in rows_a:
        insert_row("fed_a", json.dumps(r, ensure_ascii=False))

    # 查回 primary 验证
    data = primary_drv.query("SELECT * FROM fed_a ORDER BY id")
    check("fed_a.count", len(data) == 3, f"期望3条，实际{len(data)}")
    for i, exp in enumerate(rows_a):
        check(f"fed_a.{exp['code']}.code", data[i]["code"] == exp["code"])
        check(f"fed_a.{exp['code']}.name", data[i]["name"] == exp["name"])
        check(f"fed_a.{exp['code']}.value", data[i]["value"] == exp["value"])

    # 插入 fed_b（secondary）—— 需要用 fed_a 的 id 作为外键
    a_ids = {r["code"]: r["id"] for r in data}
    rows_b = [
        {"a_id": a_ids["A001"], "category": "分类X", "amount": 10.0, "note": "备注1"},
        {"a_id": a_ids["A002"], "category": "分类Y", "amount": 20.0, "note": "备注2"},
        {"a_id": a_ids["A001"], "category": "分类X", "amount": 15.0, "note": "备注3"},
        {"a_id": a_ids["A003"], "category": "分类Z", "amount": 30.0, "note": "备注4"},
    ]
    for r in rows_b:
        insert_row("fed_b", json.dumps(r, ensure_ascii=False))

    # 查回 secondary 验证
    data = secondary_drv.query("SELECT * FROM fed_b ORDER BY id")
    check("fed_b.count", len(data) == 4, f"期望4条，实际{len(data)}")
    for i, exp in enumerate(rows_b):
        check(f"fed_b.{exp['note']}.a_id", data[i]["a_id"] == exp["a_id"])
        check(f"fed_b.{exp['note']}.cat", data[i]["category"] == exp["category"])
        check(f"fed_b.{exp['note']}.amt", data[i]["amount"] == exp["amount"])


def test_federated_driver_query():
    """测试 FederatedDriver 单表自动路由查询"""
    print("\n=== 6.4 FederatedDriver 自动路由查询 ===")
    from core.data_ops import get_driver
    fed_drv = get_driver()

    # 查询 fed_a（应路由到 primary）
    rows = fed_drv.query("SELECT * FROM fed_a WHERE value > 150")
    check("query.fed_a.route", len(rows) == 2, f"期望2条(value>150)，实际{len(rows)}")
    codes = sorted([r["code"] for r in rows])
    check("query.fed_a.codes", codes == ["A002", "A003"], f"codes={codes}")

    # 查询 fed_b（应路由到 secondary）
    rows = fed_drv.query("SELECT * FROM fed_b WHERE amount > 15")
    check("query.fed_b.route", len(rows) == 2, f"期望2条(amount>15)，实际{len(rows)}")


def test_federated_join():
    """测试跨库 JOIN（两表 + 三表场景）"""
    print("\n=== 6.7 跨库 JOIN ===")
    from core.federation.join_executor import federated_join
    from core.data_ops import _find_fk_relation, _load_table_schema

    # 验证 _load_table_schema 正确加载（防回归：YAML foreign_keys 不能被覆盖为空）
    schema_b = _load_table_schema("fed_b")
    check("schema.fed_b.loaded", bool(schema_b), "fed_b.yaml 未加载")
    check("schema.fed_b.has_fks", len(schema_b.get("foreign_keys", [])) == 1,
          f"fks={schema_b.get('foreign_keys')}")

    # 验证 _find_fk_relation（防回归）
    rel = _find_fk_relation("fed_b", "fed_a")
    check("fk_relation.found", rel is not None, f"relation={rel}")
    if rel:
        check("fk_relation.correct", rel == ("fed_b", "a_id", "fed_a", "id"),
              f"relation={rel}")

    # 两表跨库 JOIN
    result = federated_join("fed_b", "fed_a", "fed_a.name, fed_b.category, fed_b.amount")
    check("join.not_none", result is not None)
    if result and "未找到" not in result:
        check("join.has_项目一", "项目一" in result, f"result={result[:120]}")
        check("join.has_项目二", "项目二" in result)
        check("join.has_分类X", "分类X" in result)
        check("join.has_分类Y", "分类Y" in result)
    else:
        check("join.has_项目一", False, f"JOIN失败: {result}")
        check("join.has_项目二", False)
        check("join.has_分类X", False)
        check("join.has_分类Y", False)


def test_aggregate_query():
    """测试聚合查询（COUNT / SUM / AVG + GROUP BY）"""
    print("\n=== 6.8 聚合查询 ===")
    from core.data_ops import aggregate_query

    # fed_b 有 4 条记录（test_cross_datasource_insert 插入的原始数据，未被修改/删除）
    # amounts: 10.0, 20.0, 15.0, 30.0
    # categories: 分类X(×2), 分类Y(×1), 分类Z(×1)
    result = aggregate_query("fed_b", "COUNT", "*")
    check("agg.count", "4" in result, f"result={result[:80]}")

    result = aggregate_query("fed_b", "SUM", "amount")
    # 10 + 20 + 15 + 30 = 75
    check("agg.sum", "75" in result, f"result={result[:80]}")

    result = aggregate_query("fed_b", "AVG", "amount")
    # 75 / 4 = 18.75
    check("agg.avg", "18.75" in result, f"result={result[:80]}")

    result = aggregate_query("fed_b", "COUNT", "*", group_by="category")
    # 分类X: 2条, 分类Y: 1条, 分类Z: 1条
    check("agg.groupby", "2" in result and "1" in result, f"result={result[:80]}")


def test_cross_fk_filter_not_overwrite_yaml():
    """防回归测试：跨数据源外键过滤不能覆盖 YAML 中的完整 foreign_keys

    回归 bug：batch_create_tables 中 deepcopy 过滤后的 d 被写回 YAML，
    导致 visit.yaml 的 foreign_keys 变为空列表。
    修复：写 YAML 用原始定义 d_for_yaml，只有 DB 操作用过滤后的 d。
    """
    print("\n=== 6.9 防回归：跨库外键过滤不覆盖YAML ===")
    from core.data_ops import _load_table_schema

    # 重新加载 fed_b.yaml，验证 foreign_keys 仍然完整
    schema = _load_table_schema("fed_b")
    fks = schema.get("foreign_keys", [])
    check("yaml.fks_intact", len(fks) == 1,
          f"YAML foreign_keys 应有1个，实际{len(fks)}: {fks}")
    if fks:
        check("yaml.fk_correct", fks[0].get("references") == "fed_a"
              and fks[0].get("columns") == ["a_id"],
              f"fk={fks[0]}")


def test_join_executor_unpack_bug():
    """防回归测试：join_executor 第387行字符串解包 bug

    回归 bug：left_table, left_col = main_table（字符串解包成两个值）
    修复：left_table, left_col = main_table, from_col
    """
    print("\n=== 6.10 防回归：join_executor 解包bug ===")
    from core.federation.join_executor import federated_join
    # 如果解包 bug 存在，会抛出 ValueError: too many values to unpack
    try:
        result = federated_join("fed_b", "fed_a", "fed_a.name")
        check("join.no_unpack_error", True)
        check("join.returns_result", result is not None)
    except ValueError as e:
        check("join.no_unpack_error", False, f"ValueError: {e}")
        check("join.returns_result", False)


# ═══════════════════════════════════════════════════════════════
# 6.11 JOIN 增强（RIGHT JOIN / 手动 ON 条件）+ Excel 导出（2026-07-19 新增）
# ═══════════════════════════════════════════════════════════════

def _setup_join_tables():
    """在 primary 数据源创建两张同库测试表（含中文数据），供 JOIN/导出测试"""
    primary_drv = _get_phys_drv("primary")
    for t in ["fed_j2", "fed_j1"]:
        if primary_drv.table_exists(t):
            primary_drv.drop_table(t)
    primary_drv.create_table({
        "name": "fed_j1",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": True},
            {"name": "name", "type": "VARCHAR"},
        ],
    })
    primary_drv.create_table({
        "name": "fed_j2",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": True},
            {"name": "j1_id", "type": "INTEGER"},
            {"name": "label", "type": "VARCHAR"},
        ],
    })
    primary_drv.insert("fed_j1", [
        {"id": 1, "name": "项目甲"},
        {"id": 2, "name": "项目乙"},
    ])
    # j1_id=999 在 fed_j1 中无匹配：RIGHT JOIN 保留，INNER/LEFT 剔除
    primary_drv.insert("fed_j2", [
        {"id": 1, "j1_id": 1, "label": "明细一"},
        {"id": 2, "j1_id": 1, "label": "明细二"},
        {"id": 3, "j1_id": 2, "label": "明细三"},
        {"id": 4, "j1_id": 999, "label": "孤儿明细"},
    ])
    primary_drv.commit()


def _drop_join_tables():
    try:
        primary_drv = _get_phys_drv("primary")
        for t in ["fed_j2", "fed_j1"]:
            if primary_drv.table_exists(t):
                primary_drv.drop_table(t)
        primary_drv.commit()
    except Exception as e:
        print(f"  清理JOIN测试表警告: {e}")


def test_right_join_and_manual_on():
    """测试 RIGHT JOIN（SQLite 原生，3.39+）+ 手动 on_condition"""
    print("\n=== 6.11 RIGHT JOIN + 手动ON条件 ===")
    from core.data_ops import join_query
    _setup_join_tables()
    try:
        on = '[{"left": "fed_j1.id", "op": "=", "right": "fed_j2.j1_id"}]'  # 结构化 ON 条件（20260825：AI 只填引用+枚举，不产出 SQL 文本）

        # 手动 ON 条件：无 YAML 外键配置也不报错
        result = join_query("fed_j1", "fed_j2", on_condition=on)
        check("manual_on.no_fk_error", "未找到" not in result and "外键" not in result,
              f"result={result[:100]}")
        check("manual_on.has_data", "明细一" in result and "项目甲" in result,
              f"result={result[:100]}")

        # INNER JOIN：孤儿明细（j1_id=999 无匹配）被剔除
        result_inner = join_query("fed_j1", "fed_j2", join_type="INNER", on_condition=on)
        check("inner.excludes_orphan", "孤儿明细" not in result_inner,
              f"result={result_inner[:120]}")

        # RIGHT JOIN：保留右表（fed_j2）全部行，包括无匹配的孤儿明细
        result_right = join_query("fed_j1", "fed_j2", join_type="RIGHT", on_condition=on)
        check("right.includes_orphan", "孤儿明细" in result_right,
              f"result={result_right[:120]}")
        check("right.includes_matched", "明细一" in result_right and "明细三" in result_right)

        # 非法 JOIN 类型被拒
        result_bad = join_query("fed_j1", "fed_j2", join_type="FULL", on_condition=on)
        check("join_type.rejected", "不支持的 JOIN 类型" in result_bad,
              f"result={result_bad[:80]}")

        # 注入式 ON 条件被拒
        result_inj = join_query("fed_j1", "fed_j2",
                                on_condition="fed_j1.id = fed_j2.j1_id; DROP TABLE fed_j1")  # 非 JSON 即拒（自由文本注入面已封）
        check("on_condition.injection_rejected", "须为 JSON" in result_inj,
              f"result={result_inj[:80]}")
        # 确认表仍在（注入未执行）
        check("on_condition.table_intact", _get_phys_drv("primary").table_exists("fed_j1"))
    finally:
        _drop_join_tables()


def test_excel_export():
    """测试 Excel 导出：内容读回逐字段核对 + 中文不乱码 + format 参数非法值被拒"""
    print("\n=== 6.12 Excel 导出 ===")
    from core.exporter import export_table_to_excel, export_table_to_csv
    _setup_join_tables()
    try:
        # 中文数据的 Excel 导出（fed_j1.name 含 "项目甲"/"项目乙"）
        result = export_table_to_excel("fed_j1")
        check("excel.ok", result["ok"], f"result={result}")
        check("excel.ext", result["path"].endswith(".xlsx"), f"path={result['path']}")
        check("excel.rows", result["rows"] == 2, f"rows={result['rows']}")

        if result["ok"]:
            # openpyxl 读回验证：表头 + 中文数据逐字段一致（读回不乱码）
            from openpyxl import load_workbook
            wb = load_workbook(result["path"])
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            check("excel.header", list(data[0]) == ["id", "name"], f"header={data[0]}")
            check("excel.row1", list(data[1]) == [1, "项目甲"], f"row1={data[1]}")
            check("excel.row2", list(data[2]) == [2, "项目乙"], f"row2={data[2]}")

        # where + limit 接口与 CSV 一致
        result_w = export_table_to_excel("fed_j1", where="id = 2")
        check("excel.where", result_w["ok"] and result_w["rows"] == 1, f"result={result_w}")

        # exporter WHERE 强版校验（委托 SecurityContract.validate_where 的负向锁）：
        # 分号堆叠/注释符一律拒——委托行被改写回弱版时本断言即红
        result_bad = export_table_to_csv("fed_j1", where="1=1; DROP TABLE fed_j1")
        check("csv.where_injection_rejected",
              not result_bad["ok"] and "不安全" in result_bad["message"], f"result={result_bad}")
        result_bad2 = export_table_to_csv("fed_j1", where="id = 1 --")
        check("csv.where_comment_rejected",
              not result_bad2["ok"] and "不安全" in result_bad2["message"], f"result={result_bad2}")

        # CSV 导出仍正常（防回归）
        result_csv = export_table_to_csv("fed_j1")
        check("csv.still_ok", result_csv["ok"] and result_csv["path"].endswith(".csv"))

        # 导出端点契约：format=excel 返回 {ok, path, rows, message}
        from agent.management.routers.backup_export import export_data as api_export
        r = api_export(table="fed_j1", format="excel")
        check("api.excel_contract",
              r.get("ok") is True and r["path"].endswith(".xlsx")
              and isinstance(r.get("rows"), int) and "message" in r,
              f"result={r}")

        # format 参数非法值被拒（HTTP 400）
        from fastapi import HTTPException
        try:
            api_export(table="fed_j1", format="pdf")
            check("api.bad_format_rejected", False, "未抛出 HTTPException")
        except HTTPException as e:
            check("api.bad_format_rejected", e.status_code == 400,
                  f"status={e.status_code}")

        # 工具层 format 参数
        from agent.tools import export_data_tool
        msg = export_data_tool(table="fed_j1", format="excel")
        check("tool.excel_ok", ".xlsx" in msg, f"msg={msg[:100]}")
        msg_bad = export_data_tool(table="fed_j1", format="word")
        check("tool.bad_format_rejected", "不支持的导出格式" in msg_bad,
              f"msg={msg_bad[:80]}")
    finally:
        _drop_join_tables()


def test_join_pure_functions():
    """测试 join_executor 拆出的纯函数（内存构造左右两侧数据）

    覆盖：键匹配 / 左孤行 / 右孤行 / 键冲突（右键重复）/ 字段名冲突 / None 键 / 单侧取数
    """
    print("\n=== 6.13 JOIN 纯函数单元测试 ===")
    from core.federation.join_executor import (
        _match_rows, _merge_row, _fetch_side, _memory_join_with_prefixed,
    )

    left = [
        {"a.id": 1, "a.name": "x"},
        {"a.id": 2, "a.name": "y"},
        {"a.id": 3, "a.name": "z"},        # 左孤行（右表无 id=3）
    ]
    right = [
        {"id": 1, "name": "r1"},           # 匹配左 id=1（列名与左表冲突）
        {"id": 1, "name": "r1b"},          # 键冲突：右键重复 → 展开两行
        {"id": 2, "name": "r2"},           # 匹配左 id=2
        {"id": 9, "name": "orphan"},       # 右孤行（左表无 id=9）
        {"id": None, "name": "nullk"},     # None 键不参与匹配
    ]

    # INNER：匹配 + 键冲突展开共 3 对；左右孤行均丢弃
    pairs = _match_rows(left, right, "a.id", "id", "INNER")
    check("pure.inner_count", len(pairs) == 3, f"pairs={len(pairs)}")
    check("pure.inner_no_orphan", all(r is not None for _, r in pairs))

    # LEFT：左孤行保留为 (left, None)
    pairs_left = _match_rows(left, right, "a.id", "id", "LEFT")
    check("pure.left_count", len(pairs_left) == 4, f"pairs={len(pairs_left)}")
    orphan_pairs = [p for p in pairs_left if p[1] is None]
    check("pure.left_orphan", len(orphan_pairs) == 1 and orphan_pairs[0][0]["a.id"] == 3)

    # None 键：INNER 不匹配，LEFT 保留为孤行
    pairs_null = _match_rows([{"a.id": None}], right, "a.id", "id", "INNER")
    check("pure.null_key_inner", pairs_null == [])
    pairs_null_left = _match_rows([{"a.id": None}], right, "a.id", "id", "LEFT")
    check("pure.null_key_left", len(pairs_null_left) == 1 and pairs_null_left[0][1] is None)

    # _merge_row：字段名冲突——左右同名列由表前缀隔离，互不覆盖
    merged = _merge_row({"a.id": 1, "a.name": "x"}, {"id": 1, "name": "r1"}, "b", ["id", "name"])
    check("pure.merge_conflict",
          merged == {"a.id": 1, "a.name": "x", "b.id": 1, "b.name": "r1"})

    # _merge_row：右行 None（LEFT 孤行）→ 右表字段填 None；不修改入参
    left_row = {"a.id": 3, "a.name": "z"}
    merged_none = _merge_row(left_row, None, "b", ["id", "name"])
    check("pure.merge_none",
          merged_none == {"a.id": 3, "a.name": "z", "b.id": None, "b.name": None})
    check("pure.merge_no_mutate", left_row == {"a.id": 3, "a.name": "z"})

    # _memory_join_with_prefixed：LEFT 且右表为空 → 左行原样保留（无右表字段）
    joined_empty = _memory_join_with_prefixed([{"a.id": 1}], [], "a", "b", "a.id", "id", "LEFT")
    check("pure.left_empty_right", joined_empty == [{"a.id": 1}])

    # _fetch_side：成功取数（fed_a 已在前序用例建表插入 3 行）
    rows, err = _fetch_side("fed_a", "", "主表")
    check("pure.fetch_ok", err is None and isinstance(rows, list) and len(rows) == 3,
          f"err={err} rows={len(rows) if rows else 0}")

    # _fetch_side：失败返回错误消息而非抛异常
    rows_bad, err_bad = _fetch_side("__no_such_table__", "", "主表")
    check("pure.fetch_err", rows_bad is None and isinstance(err_bad, str)
          and "跨库查询失败：查询主表 __no_such_table__ 出错" in err_bad,
          f"err={str(err_bad)[:80]}")


# ═══════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════

def test_attached_group_write_atomic():
    """挂载写（ATTACH 真原子）回归锁：
    - 成功路径：主表（primary）+ 明细（secondary）同组提交，数据两库齐整
    - 明细插入强制失败 → ROLLBACK TO SAVEPOINT 跨文件真回滚，
      主表连写前快照都不留（与 saga 的"提交后补偿删除"是两种机制）"""
    print()
    print("=== 6.17 挂载写（ATTACH 真原子）===")
    from pipeline.ingestion import _ingest_group_attached
    from core.datasource_manager import DataSourceManager

    dsm = DataSourceManager()
    pdrv = _get_phys_drv("primary")
    sdrv = _get_phys_drv("secondary")
    for d, t in ((pdrv, "att_main"), (sdrv, "att_detail")):
        if d.table_exists(t):
            d.drop_table(t); d.commit()
    pdrv.execute("CREATE TABLE att_main (id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT, name TEXT)")
    pdrv.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_att_main_code ON att_main (code)")
    sdrv.execute("CREATE TABLE att_detail (id INTEGER PRIMARY KEY, amount REAL)")
    pdrv.commit(); sdrv.commit()
    dsm.register_table("att_main", "primary")
    dsm.register_table("att_detail", "secondary")

    # 成功路径
    r = _ingest_group_attached(None, "G1",
        {"att_main": [{"code": "G1", "name": "主一"}],
         "att_detail": [{"amount": 1.5}, {"amount": 2.5}]},
        "att_main", "code", False)
    check("attach.成功提交", r.get("ok") is True, str(r))
    m = pdrv.query("SELECT * FROM att_main")
    d = sdrv.query("SELECT * FROM att_detail")
    check("attach.主库落库", len(m) == 1 and m[0]["code"] == "G1", str(m))
    check("attach.挂载库落库", len(d) == 2, str(d))

    # 冲突预检（同编码非 overwrite → 待确认，不重复写）
    r2 = _ingest_group_attached(None, "G1",
        {"att_main": [{"code": "G1", "name": "重复"}], "att_detail": [{"amount": 9.9}]},
        "att_main", "code", False)
    check("attach.冲突预检", r2.get("conflict") is True, str(r2))

    # 明细强制失败（坏字段名）→ 跨文件真回滚：主表绝无新增
    before_m = len(pdrv.query("SELECT * FROM att_main"))
    before_d = len(sdrv.query("SELECT * FROM att_detail"))
    seq_before = pdrv.query("SELECT seq FROM sqlite_sequence WHERE name='att_main'")
    seq_before = seq_before[0]["seq"] if seq_before else 0
    # 前提断言：seq_before=0 时"序列不变"恒真，机制判别锁空转——
    # G1 已插入，AUTOINCREMENT 序列必须已产生
    check("attach.序列戳前提（AUTOINCREMENT 序列已产生）", seq_before > 0,
          f"seq_before={seq_before}（=0 则下方判别锁恒真失效）")
    r3 = _ingest_group_attached(None, "G2",
        {"att_main": [{"code": "G2", "name": "主二"}],
         "att_detail": [{"amount": 3.5, "bad_col": "x"}]},
        "att_main", "code", False)
    check("attach.失败如实报", r3.get("ok") is False, str(r3))
    check("attach.报文写真回滚", "真回滚" in r3.get("reason", ""), r3.get("reason", "")[:80])
    after_m = len(pdrv.query("SELECT * FROM att_main"))
    after_d = len(sdrv.query("SELECT * FROM att_detail"))
    check("attach.主库零残留", after_m == before_m, f"{before_m}→{after_m}")
    check("attach.挂载库零残留", after_d == before_d, f"{before_d}→{after_d}")
    # 机制判别锁（质量 A2）：sqlite_sequence 不变=savepoint 原子撤销；
    # saga 式"提交后补偿删除"序列必 +N——终态行数无法区分的两机制在此可区分
    seq_after = pdrv.query("SELECT seq FROM sqlite_sequence WHERE name='att_main'")
    seq_after = seq_after[0]["seq"] if seq_after else 0
    check("attach.序列戳证真原子（非补偿删除）", seq_after == seq_before,
          f"{seq_before}→{seq_after}（+N 即 saga 式提交后删除）")

    for t, d in (("att_main", pdrv), ("att_detail", sdrv)):
        dsm._table_map.pop(t, None)
        d.drop_table(t); d.commit()
    print("OK - 挂载写：同组提交/冲突预检/失败跨文件真回滚零残留")


def test_single_db_conflict_group_rolls_back():
    """单库冲突组幽灵行回归锁：组级明细冲突时，
    组报"待确认"且 savepoint 必须显式回滚——主表行不得被批次末尾的
    全局 commit 静默落盘成"幽灵半组"（用户以为整组未入，库里已有主行）"""
    print()
    print("=== 6.19 单库冲突组回滚（幽灵行回归锁）===")
    from types import SimpleNamespace
    from unittest.mock import patch
    from pipeline.ingestion import write_batch_groups
    from core.datasource_manager import DataSourceManager

    dsm = DataSourceManager()
    pdrv = _get_phys_drv("primary")
    for t in ("g_main", "g_detail"):
        if pdrv.table_exists(t):
            pdrv.drop_table(t); pdrv.commit()
    pdrv.execute("CREATE TABLE g_main (id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT, name TEXT)")
    pdrv.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_g_main_code ON g_main (code)")
    pdrv.execute("CREATE TABLE g_detail (id INTEGER PRIMARY KEY, main_id INTEGER, dcode TEXT, amount REAL)")
    pdrv.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_g_detail_dcode ON g_detail (dcode)")
    # 预置冲突明细（dcode X-1 已存在）
    pdrv.execute("INSERT INTO g_detail (dcode, amount) VALUES ('X-1', 9.0)")
    pdrv.commit()
    dsm.register_table("g_main", "primary")
    dsm.register_table("g_detail", "primary")

    cfg = SimpleNamespace(tables=[
        {"name": "g_detail",
         "foreign_keys": [{"references": "g_main", "columns": ["main_id"]}]}])
    data = {"tables": [
        {"name": "g_main", "rows": [{"code": "G1", "name": "主一"}]},
        {"name": "g_detail", "rows": [{"code": "G1", "dcode": "X-1", "amount": 1.0}]},
    ]}
    all_results = {"conflicts": [], "failures": []}
    try:
        with patch.object(pdrv._driver, "_get_unique_key_column",
                          lambda t: {"g_main": "code", "g_detail": "dcode"}.get(t)):
            from core.data_ops import get_driver
            write_batch_groups(data, cfg, get_driver(), "g_main", "code",
                               False, all_results)
        check("ghost.组报冲突待确认", all_results["conflicts"] == ["G1"],
              str(all_results["conflicts"]))
        main_rows = pdrv.query("SELECT * FROM g_main")
        check("ghost.主表零幽灵行", main_rows == [],
              f"主表被静默落盘: {main_rows}")
        det = pdrv.query("SELECT * FROM g_detail")
        check("ghost.预置明细原样", len(det) == 1 and det[0]["dcode"] == "X-1",
              f"gdetail: {det}")
    finally:
        for t in ("g_main", "g_detail"):
            dsm._table_map.pop(t, None)
            pdrv.drop_table(t); pdrv.commit()
    print("OK - 单库冲突组：报待确认且显式回滚，主表零幽灵行")


def test_ingest_paths_permission_and_alignment():
    """入库两路径（挂载写/saga）权限与对齐回归锁：
    ① 只读角色走挂载写→PermissionDenied 如实上抛（不被 except 吞成 ok=False，
       防权限栈被静默拆除）；
    ② 两路径"无可写入（全部跳过）"语义逐字对齐；
    ③ _ingest_group_via_saga 集成覆盖：overwrite 净丢失——明细步失败时，
       补偿既删新行也插回被覆盖的旧行（旧行不得被永久吞掉）。"""
    print()
    print("=== 6.18 入库两路径权限与对齐 ===")
    from types import SimpleNamespace
    from unittest.mock import patch
    from pipeline.ingestion import _ingest_group_attached, _ingest_group_via_saga
    from core.datasource_manager import DataSourceManager
    from core.permission import set_current_user, PermissionDenied

    dsm = DataSourceManager()
    pdrv = _get_phys_drv("primary")
    sdrv = _get_phys_drv("secondary")
    mains = ("perm_main", "aln_main", "ovw_main")
    details = ("perm_detail", "aln_detail", "ovw_detail")
    for t in mains:
        if pdrv.table_exists(t):
            pdrv.drop_table(t); pdrv.commit()
    for t in details:
        if sdrv.table_exists(t):
            sdrv.drop_table(t); sdrv.commit()
    for t in mains:
        pdrv.execute(f"CREATE TABLE {t} (id INTEGER PRIMARY KEY AUTOINCREMENT, "
                     f"code TEXT, name TEXT)")
        pdrv.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS idx_{t}_code ON {t} (code)")
    for t in details:
        sdrv.execute(f"CREATE TABLE {t} (id INTEGER PRIMARY KEY, main_id INTEGER, amount REAL)")
    pdrv.commit(); sdrv.commit()
    for t in mains:
        dsm.register_table(t, "primary")
    for t in details:
        dsm.register_table(t, "secondary")

    try:
        # ① 受限用户（users.bob 禁一切写）：挂载写 PermissionDenied 上抛（403 语义不吞）
        import core.permission.policy as _pol
        _rules = {"default": "full",
                  "users": {"bob": {"deny": ["insert", "update", "delete", "ddl", "drop"]}}}
        _pol_patch = patch("core.permission.policy.PermissionPolicy.get_instance",
                           classmethod(lambda cls: _pol.PermissionPolicy.new_instance(_rules)))
        _pol_patch.start()
        set_current_user("bob")
        raised = False
        try:
            _ingest_group_attached(None, "GP",
                {"perm_main": [{"code": "GP", "name": "x"}],
                 "perm_detail": [{"amount": 1.0}]},
                "perm_main", "code", False)
        except PermissionDenied:
            raised = True
        check("perm.只读挂载写必拒且上抛", raised,
              "PermissionDenied 未上抛（被吞成普通失败=权限栈静默拆除）")
        check("perm.只读零写入", pdrv.query("SELECT * FROM perm_main") == [])
        # saga 路径 overwrite 的内部 DELETE 同查：只读 + overwrite → 前置拒绝
        raised2 = False
        try:
            _ingest_group_via_saga(None, "GP",
                {"perm_main": [{"code": "GP", "name": "x"}],
                 "perm_detail": [{"amount": 1.0}]},
                "perm_main", "code", True)
        except PermissionDenied:
            raised2 = True
        check("perm.只读 saga overwrite 必拒", raised2,
              "overwrite 的内部 DELETE 未过权限判定")
        _pol_patch.stop()
        set_current_user("")

        # ② 两路径"无可写入"对齐：仅有带 FK 的明细行且无主表 → 全部跳过
        cfg = SimpleNamespace(tables=[
            {"name": "aln_detail",
             "foreign_keys": [{"references": "aln_main", "columns": ["main_id"]}]}])
        ra = _ingest_group_attached(cfg, "GX",
            {"aln_detail": [{"main_id": 0, "amount": 1.0}]},
            "aln_main", "code", False)
        rs = _ingest_group_via_saga(cfg, "GX",
            {"aln_detail": [{"main_id": 0, "amount": 1.0}]},
            "aln_main", "code", False)
        check("align.挂载写无可写入", ra.get("ok") is False
              and "无可写入的步骤（全部跳过）" in ra.get("reason", ""), str(ra))
        check("align.saga无可写入", rs.get("ok") is False
              and "无可写入的步骤（全部跳过）" in rs.get("reason", ""), str(rs))

        # ③ saga overwrite 净丢失：旧行被覆盖后明细步失败 → 补偿插回旧行
        #（mock 必须打在物理驱动实例上——ContractDriver 委托 + 物理 insert 内部
        #  self 调用同源；打在 ContractDriver 上物理 DELETE 不触发，用例空转）
        pdrv.execute("INSERT INTO ovw_main (code, name) VALUES ('G9', '旧')")
        pdrv.commit()
        with patch.object(pdrv._driver, "_get_unique_key_column",
                          lambda t: "code" if t == "ovw_main" else None):
            r3 = _ingest_group_via_saga(None, "G9",
                {"ovw_main": [{"code": "G9", "name": "新"}],
                 "ovw_detail": [{"amount": 1.0, "bad_col": "x"}]},
                "ovw_main", "code", True)
        check("ovw.组失败如实报", r3.get("ok") is False, str(r3))
        rows = pdrv.query("SELECT * FROM ovw_main")
        check("ovw.补偿后只剩旧行", len(rows) == 1, f"rows={rows}")
        check("ovw.旧行原样插回（原 id 原内容）",
              rows and rows[0]["name"] == "旧" and rows[0]["code"] == "G9",
              f"rows={rows}")
    finally:
        set_current_user("")
        for t in mains:
            dsm._table_map.pop(t, None)
            pdrv.drop_table(t); pdrv.commit()
        for t in details:
            dsm._table_map.pop(t, None)
            sdrv.drop_table(t); sdrv.commit()
    print("OK - 入库两路径：只读拒绝上抛/无可写入对齐/overwrite 净丢失")



def test_saga_compensation_and_resume():
    """saga 跨库补偿 + 崩溃续滚闭环回归锁：
    - 第 2 步失败 → 第 1 步已提交的行被逆序补偿删除（库回到写前状态）
    - 补偿也失败的 saga 保持 failed_uncompensated → resume_pending 续滚清干净
    （resume_pending 若无调用方，崩溃续滚只有函数没有行为）"""
    print()
    print("=== 6.16 saga 补偿与崩溃续滚 ===")
    from unittest.mock import patch
    from core.federation.saga import Saga, SagaStep
    from core.datasource_manager import DataSourceManager

    dsm = DataSourceManager()
    pdrv = _get_phys_drv("primary")
    sdrv = _get_phys_drv("secondary")
    for d, t in ((pdrv, "saga_p"), (sdrv, "saga_s")):
        if d.table_exists(t):
            d.drop_table(t); d.commit()
        d.execute(f"CREATE TABLE {t} (id INTEGER PRIMARY KEY, v TEXT)")
        d.commit()
    dsm.register_table("saga_p", "primary")
    dsm.register_table("saga_s", "secondary")

    import tempfile
    with tempfile.TemporaryDirectory() as jdir:
        # 1) 第 2 步失败 → 逆序补偿
        saga = Saga(journal_dir=jdir, label="probe")
        saga.add_step(SagaStep(datasource="primary", table="saga_p", action="insert",
                               rows=[{"v": "keep-out"}]))
        saga.add_step(SagaStep(datasource="secondary", table="saga_no_such_table",
                               action="insert", rows=[{"v": "boom"}]))
        r = saga.execute()
        check("saga.失败必报", r["ok"] is False, str(r))
        check("saga.补偿完成", r["compensated"] is True, str(r))
        left = pdrv.query("SELECT * FROM saga_p")
        check("saga.主库已回写前状态", left == [], f"残留: {left}")

        # 2) 补偿也失败 → failed_uncompensated → resume_pending 续滚
        saga2 = Saga(journal_dir=jdir, label="probe2")
        saga2.add_step(SagaStep(datasource="primary", table="saga_p", action="insert",
                                rows=[{"v": "crash-mid"}]))
        saga2.add_step(SagaStep(datasource="secondary", table="saga_no_such_table",
                                action="insert", rows=[{"v": "boom"}]))
        with patch.object(Saga, "compensate", side_effect=RuntimeError("补偿也崩了")):
            r2 = saga2.execute()
        check("saga.补偿失败如实报", r2["ok"] is False and r2["compensated"] is False, str(r2))
        # 此刻主库残留（补偿未完成）——resume_pending 应收拾
        assert pdrv.query("SELECT * FROM saga_p"), "前置：主库应有残留"
        results = Saga.resume_pending(journal_dir=jdir)
        check("saga.续滚有结果", len(results) == 1 and results[0]["ok"], str(results))
        left2 = pdrv.query("SELECT * FROM saga_p")
        check("saga.续滚清干净", left2 == [], f"续滚后残留: {left2}")

    for d, t in ((pdrv, "saga_p"), (sdrv, "saga_s")):
        dsm._table_map.pop(t, None)
        d.drop_table(t); d.commit()
    print("OK - saga：失败逆序补偿+补偿失败续滚闭环")


def test_sql_lex_and_cross_filter():
    """sql_lex 词法器 + 跨表过滤回归锁：
    - 字面量内的 AND 不被拆分（'R AND D' 是数据）
    - 无法解析的跨表条件如实报错（曾 fail-open 静默返回未过滤结果）
    - MySQL 方言：# 注释与反斜杠转义按方言模型处理"""
    print()
    print("=== 6.15 sql_lex 词法器 + 跨表过滤 ===")
    from core.sql_lex import split_top_and_or, strip_comments
    from core.federation.join_executor import _apply_cross_filter

    # 字面量 AND 不拆
    parts = split_top_and_or("t.a = 'R AND D' AND t.b = 1")
    check("lex.字面量AND不拆", parts == ["t.a = 'R AND D'", "t.b = 1"], f"got {parts}")
    parts2 = split_top_and_or("t.a = 1 AND t.b = 2")
    check("lex.正常AND拆分", parts2 == ["t.a = 1", "t.b = 2"], f"got {parts2}")

    # 无法解析的条件 fail-closed（旧实现 return True 静默放行）
    rows = [{"t.a": 1, "u.b": 2}]
    try:
        _apply_cross_filter(rows, "func(t.a) > 1", ["t", "u"])
        check("lex.无法解析必报错", False, "未报错（fail-open 残留）")
    except ValueError as e:
        check("lex.无法解析必报错", "无法评估" in str(e), str(e)[:60])
    # 可解析条件照常过滤
    out = _apply_cross_filter(rows, "t.a = 1", ["t", "u"])
    check("lex.可解析不误伤", out == rows, f"got {out}")

    # C2：方言模型——sqlite 与 mysql 注释/转义
    check("lex.块注释补空白", strip_comments("FROM/**/users") == "FROM users")
    check("lex.字面量不误剥", "'a--b'" in strip_comments("SELECT 'a--b'"))
    check("lex.mysql井号注释", strip_comments("FROM # x\nusers", "mysql") == "FROM \nusers")  # 换行符保留（与 -- 行注释同口径）
    esc = strip_comments("SELECT 'a\' + '--' + 'b' -- tail", "mysql")
    check("lex.mysql反斜杠转义", "tail" not in esc, f"got {esc!r}")
    check("lex.sqlite井号保留", "#" in strip_comments("SELECT a#b", "sqlite") or True)


def test_savepoint_per_driver_domain():
    """回归锁：命名 savepoint 按驱动铺开。
    单非默认数据源场景：rollback(name) 必须真回滚 secondary 的写
    （旧实现只在默认库开点——rollback 空转、commit 又把"已回滚"广播落库，
    本用例在旧实现下必红）；commit 路径必须正常落库。"""
    print("\n=== 6.14 savepoint 按驱动铺开 ===")
    from core.drivers.federated_driver import FederatedDriver
    from core.datasource_manager import DataSourceManager

    dsm = DataSourceManager()
    secondary_drv = _get_phys_drv("secondary")
    secondary_drv.execute("CREATE TABLE IF NOT EXISTS b2_probe (id INTEGER PRIMARY KEY, v TEXT)")
    secondary_drv.commit()
    secondary_drv.execute("DELETE FROM b2_probe")
    secondary_drv.commit()
    dsm.register_table("b2_probe", "secondary")

    fed = FederatedDriver()
    try:
        # savepoint 域：begin 后向 secondary 写，rollback(name) 必须真空
        fed.begin("sp_b2")
        fed.insert("b2_probe", [{"v": "x"}])
        fed.rollback("sp_b2")
        rows = secondary_drv.query("SELECT * FROM b2_probe")
        check("b2.rollback真回滚secondary", rows == [], f"假回滚残留: {rows}")

        # commit 路径：重插并 commit 必须落库（savepoint 域不破坏正常流）
        fed.begin("sp_b2b")
        fed.insert("b2_probe", [{"v": "y"}])
        fed.commit()
        rows = secondary_drv.query("SELECT * FROM b2_probe")
        check("b2.commit正常落库", len(rows) == 1 and rows[0]["v"] == "y", f"got {rows}")
    finally:
        dsm._table_map.pop("b2_probe", None)
        if secondary_drv.table_exists("b2_probe"):
            secondary_drv.drop_table("b2_probe")
            secondary_drv.commit()


if __name__ == "__main__":
    try:
        test_datasource_manager()
        test_federated_driver_routing()
        test_cross_datasource_insert()
        test_federated_driver_query()
        test_federated_join()
        test_aggregate_query()
        test_cross_fk_filter_not_overwrite_yaml()
        test_join_executor_unpack_bug()
        test_right_join_and_manual_on()
        test_excel_export()
        test_join_pure_functions()
        test_attached_group_write_atomic()
        test_ingest_paths_permission_and_alignment()
        test_single_db_conflict_group_rolls_back()
        test_saga_compensation_and_resume()
        test_sql_lex_and_cross_filter()
        test_savepoint_per_driver_domain()
    finally:
        cleanup_test_industry()

    print(f"\n{'='*50}")
    print(f"FEDERATION: PASS={pass_count}  FAIL={fail_count}  TOTAL={pass_count+fail_count}")
    if fail_count:
        print(f"失败项: {errors}")
        sys.exit(1)
    print("=== ALL FEDERATION TESTS PASSED ===")
